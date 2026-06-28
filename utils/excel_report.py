import os

from datetime import (
    date,
    timedelta
)

from openpyxl import Workbook

from openpyxl.styles import Font

from models.gate_pass import GatePass

from models.gate_pass_item import GatePassItem
from openpyxl.utils import get_column_letter


def generate_gatepass_excel(): 
    report_folder = "reports"

    os.makedirs(
        report_folder,
        exist_ok=True
    )

    excel_path = os.path.join(
        report_folder,
        "gatepass_last_30_days.xlsx"
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Gate Pass Report"

    headings = [

        "Gate Pass No",

        "Created By",

        "Receiver Name",

        "Material Details"

    ]

    sheet.append(headings)

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

    today = date.today()

    last_30_days = (
        today -
        timedelta(days=30)
    ).isoformat()

    gate_passes = (
        

        GatePass.query

        .filter(
            GatePass.gatepass_date >= last_30_days
        )

        .order_by(
            GatePass.gatepass_date.desc()
        )

        .all()

    )

    for gp in gate_passes:

        items = (
            GatePassItem.query
            .filter_by(
                gate_pass_id=gp.id
            )
            .order_by(
                GatePassItem.item_no
            )
            .all()
        )

        material_details = ""

        for item in items:

            material_details += (

                f"• {item.material_description}"

                f"  | Qty : {item.qty}\n"

            )

        sheet.append(

            [

                gp.gp_number,

                gp.created_by_name,

                gp.receiver_name,

                material_details.strip()

            ]

        )

        
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 25
        sheet.column_dimensions["D"].width = 60

        for row in range(2, sheet.max_row + 1):

            sheet.row_dimensions[row].height = 60



 

    workbook.save(excel_path)

    return excel_path